"""
code_generator.py – Stok kodu bazlı P-code üretim motoru
ProfiDO (KSB_ProfilKesim) Programı
"""

import json
import os
import copy
import paths

# ── Kütüphane dosyası: sabit konumda (~/.ksb_profil/) ──────────────────────
# Dahili format: JSON (hızlı). Excel sadece dışa/içe aktarım için kullanılır.
_LIB_DIR  = paths.app_data_dir()
_LIB_PATH = os.path.join(_LIB_DIR, 'profile_library.json')
_XLS_PATH = os.path.join(_LIB_DIR, 'profile_library.xlsx')   # dışa aktarım
os.makedirs(_LIB_DIR, exist_ok=True)

# Eski konum: prg dizinindeki .json → ~/.ksb_profil/ altına taşı
_OLD_JSON = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'profile_library.json')
if os.path.exists(_OLD_JSON) and not os.path.exists(_LIB_PATH):
    import shutil
    try:
        shutil.move(_OLD_JSON, _LIB_PATH)
    except Exception:
        pass
# ─────────────────────────────────────────────────────────────────────────────

_PARAM_KEYS = ['L', 'W', 'C', 'R', 'D']

# ─────────────────────────────────────────────────────────────────
# panel_tools.py ve panel_operations.py için geriye dönük uyumluluk
# ─────────────────────────────────────────────────────────────────

TOOLS = {
    'T10': {'dia': 5,          'tip': 'Freze',         'len': 25},
    'T11': {'dia': 16,         'tip': 'Freze',         'len': 25},
    'T20': {'dia': 6,          'tip': 'Matkap',        'len': 30},
    'T30': {'dia': 5,          'tip': 'Matkap',        'len': 30},
    'T31': {'dia': 18,         'tip': 'Matkap',        'len': 30},
    'T32': {'dia': 10,         'tip': 'Matkap',        'len': 30},
    'T40': {'dia': 5,          'tip': 'Freze Kasa',    'len': 25},
    'T50': {'dia': 5,          'tip': 'Matkap Montaj', 'len': 60},
    'T60': {'dia': 5,          'tip': 'Freze Kanat',   'len': 25},
    'T70': {'dia': 5,          'tip': 'Matkap Dis',    'len': 30},
    'T71': {'dia': '10-13-10', 'tip': 'Freze Kol',     'len': 20},
}

OPERATION_DESC = {
    'P1': 'Dikdörtgen Freze Kanalı',
    'P2': 'Oval Freze Kanalı',
    'P3': 'Su Tahliye Deliği',
    'P4': 'Sol Barel (Left KeyHole)',
    'P5': 'Sağ Barel (Right KeyHole)',
    'P6': 'Dairesel Delik',
    'P7': 'Nokta Deliği',
}

_OP_PARAMS = {
    'P1': [{'key': 'L', 'label': 'Uzunluk (L)'},
           {'key': 'W', 'label': 'Genişlik (W)'},
           {'key': 'D', 'label': 'Derinlik (D)'}],
    'P2': [{'key': 'L', 'label': 'Uzunluk (L)'},
           {'key': 'W', 'label': 'Genişlik (W)'},
           {'key': 'R', 'label': 'Yarıçap (R)'},
           {'key': 'D', 'label': 'Derinlik (D)'}],
    'P3': [{'key': 'L', 'label': 'Uzunluk (L)'},
           {'key': 'D', 'label': 'Derinlik (D)'}],
    'P4': [{'key': 'L', 'label': 'Uzunluk (L)'},
           {'key': 'W', 'label': 'Genişlik (W)'},
           {'key': 'C', 'label': 'Köşe (C)'},
           {'key': 'R', 'label': 'Yarıçap (R)'},
           {'key': 'D', 'label': 'Derinlik (D)'}],
    'P5': [{'key': 'L', 'label': 'Uzunluk (L)'},
           {'key': 'W', 'label': 'Genişlik (W)'},
           {'key': 'C', 'label': 'Köşe (C)'},
           {'key': 'R', 'label': 'Yarıçap (R)'},
           {'key': 'D', 'label': 'Derinlik (D)'}],
    'P6': [{'key': 'C', 'label': 'Çap (C)'},
           {'key': 'D', 'label': 'Derinlik (D)'}],
    'P7': [{'key': 'D', 'label': 'Derinlik (D)'}],
}


def get_operation_params(op: str) -> list:
    """Verilen işlem kodu için parametre tanım listesi döndürür."""
    return _OP_PARAMS.get(op, [])


def _resolve_param(params: dict, key: str) -> int:
    """
    Parametre değerini int olarak çöz.
    Sayısal değerler doğrudan; 'W/2' gibi string formüller W değeri ile hesaplanır.
    """
    v = params.get(key, 0)
    if isinstance(v, str) and v.strip():
        try:
            W = float(params.get('W', 0))
            L = float(params.get('L', 0))
            result = eval(v.strip(), {'W': W, 'L': L, '__builtins__': {}})
            return int(round(float(result)))
        except Exception:
            return 0
    try:
        return int(round(float(v))) if v else 0
    except (TypeError, ValueError):
        return 0


def build_code(operation: str, tool: str, x: float, y: float, z: float,
               params: dict) -> str:
    """
    Manuel P-code oluşturur (panel_operations.py için).
    x, y, z ve params değerleri zaten ×10 cinsinden verilir.
    R parametresi 'W/2' gibi string formül olabilir.
    """
    xi = int(round(x)); yi = int(round(y)); zi = int(round(z))
    base = f'{operation}{tool}X{xi}Y{yi}Z{zi}'

    def p(k): return _resolve_param(params, k)

    if operation == 'P7':
        return f'{base}D{p("D")}//'
    if operation == 'P3':
        if 'L' in params:
            return f'{base}L{p("L")}D{p("D")}//'
        return f'{base}D{p("D")}//'
    if operation == 'P2':
        R = p('R') if 'R' in params else p('W') // 2
        return f'{base}L{p("L")}W{p("W")}R{R}D{p("D")}//'
    if operation == 'P1':
        return f'{base}L{p("L")}W{p("W")}D{p("D")}//'
    if operation in ('P4', 'P5'):
        R = p('R') if 'R' in params else p('W') // 2
        return f'{base}L{p("L")}W{p("W")}C{p("C")}R{R}D{p("D")}//'
    if operation == 'P6':
        return f'{base}C{p("C")}D{p("D")}//'
    return f'{base}//'

SIDES = ['ALT', 'ÜST', 'SOL', 'SAĞ']

PROFILE_LABEL = {
    'A': 'Kasa',      'B': 'Kanat',    'C': 'Damlalıklı Kanat',
    'D': 'Dış Açılım','E': 'Orta Kayıt','F': 'Sürme Kasa',
    'G': 'Sürme Kanat','H': 'Pervazlı Kasa','I': 'Denizlikli Kasa',
    'J': 'Kapı Kanat',
}

# ─────────────────────────────────────────────────────────────────
# Kütüphane yükleme / kaydetme
# ─────────────────────────────────────────────────────────────────

def _empty_library() -> dict:
    return {
        '_version': 1,
        'last_used': {t: '' for t in 'ABCDEFGHIJ'},
        'tools': {},
        'profiles': {},
        'custom_ops': [],
    }


def _migrate_sides(lib: dict):
    """Eski ASCII kenar anahtarlarini (UST/SAG) Turkce'ye donustur."""
    _MAP = {'UST': 'ÜST', 'SAG': 'SAĞ'}
    for prof in lib.get('profiles', {}).values():
        ops = prof.get('operations', {})
        for old_k, new_k in _MAP.items():
            if old_k in ops:
                if not ops.get(new_k):
                    ops[new_k] = ops.pop(old_k)
                else:
                    ops.pop(old_k)


def _migrate_clean_op_names(lib: dict):
    """Makro isimlerindeki kümülatif 🔗 grup işaretlerini temizler.

    Eski bir hata: _save_ops_to_profile itm.text() ile 🔗 prefix'ini de
    kaydediyordu; her kayıtta birikirdi. Bu fonksiyon yükleme sırasında
    mevcut kütüphaneden tüm bu karakterleri tek seferde siler.
    """
    _LINK = '\U0001f517'   # 🔗
    changed = False
    for prof in lib.get('profiles', {}).values():
        if not isinstance(prof, dict):
            continue
        for macros in prof.get('operations', {}).values():
            for macro in macros:
                name = macro.get('name', '')
                if _LINK in name:
                    # replace() kullan: "🔗 🔗 🔗 Ad" gibi aralıklı durumları da temizler
                    cleaned = name.replace(_LINK, '').strip()
                    macro['name'] = cleaned
                    changed = True
    return changed


# ── Dahili hızlı format: JSON ─────────────────────────────────────────────────

def load_library(path: str = None) -> dict:
    """JSON'dan kütüphaneyi yükle (hızlı, openpyxl gerektirmez)."""
    fpath = path or _LIB_PATH
    if os.path.exists(fpath):
        try:
            with open(fpath, 'r', encoding='utf-8') as f:
                lib = json.load(f)
            lib.setdefault('custom_ops', [])
            _migrate_sides(lib)
            # 🔗 birikme temizliği: değiştiyse kütüphaneyi anında kaydet
            if _migrate_clean_op_names(lib):
                try:
                    save_library(lib, fpath)
                except Exception:
                    pass
            return lib
        except Exception:
            pass
    return _empty_library()


def save_library(library: dict, path: str = None):
    """JSON olarak kaydet (hızlı, openpyxl gerektirmez)."""
    fpath = path or _LIB_PATH
    os.makedirs(os.path.dirname(fpath), exist_ok=True)
    with open(fpath, 'w', encoding='utf-8') as f:
        json.dump(library, f, ensure_ascii=False, indent=2)


# ── Excel dışa / içe aktarım (openpyxl sadece burada yüklenir) ───────────────

def _ensure_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        import subprocess, sys
        subprocess.check_call(
            [sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        import openpyxl
        return openpyxl


def export_to_excel(library: dict, path: str = None) -> str:
    """
    Kütüphaneyi Excel dosyasına aktar.
    path verilmezse ~/.ksb_profil/profile_library.xlsx kullanılır.
    Kaydedilen dosya yolunu döner.
    """
    fpath = path or _XLS_PATH
    openpyxl = _ensure_openpyxl()
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    hdr_font  = Font(bold=True, color='FFFFFFFF')
    hdr_fill  = PatternFill('solid', start_color='FF1F4E79')
    hdr_align = Alignment(horizontal='center')

    def _hdr(ws, titles):
        ws.append(titles)
        for cell in ws[1]:
            cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align

    # Profiller
    ws_p = wb.active; ws_p.title = 'Profiller'
    _hdr(ws_p, ['Stok Kodu','Ad','Üretici','Seri','Tip','Kerf (mm)','Renk','DXF Dosyası',
                'Robot Y (×10)','Robot Z (×10)','Robot Pos'])
    for w, c in zip([18,26,16,10,8,10,12,40,12,12,10], 'ABCDEFGHIJK'):
        ws_p.column_dimensions[c].width = w
    for code, prof in library.get('profiles', {}).items():
        ws_p.append([code, prof.get('name',''), prof.get('manufacturer',''),
                     prof.get('series',''), prof.get('type','A'), prof.get('kerf',0),
                     prof.get('color','#808080'), prof.get('dxf_file',''),
                     prof.get('robot_y', 400), prof.get('robot_z', 400),
                     prof.get('robot_vertical', 0)])

    # Islemler
    ws_i = wb.create_sheet('Islemler')
    _hdr(ws_i, ['Stok Kodu','Kenar','Makro ID','Makro Adı','Adım No',
                'P-Kod','Takım','X Formülü','Y (mm)','Z (mm)',
                'L','W','C','R','D','Koşullu (JSON)'])
    for w, c in zip([18,8,22,22,8,8,8,16,8,8,6,6,6,6,6,35], 'ABCDEFGHIJKLMNOP'):
        ws_i.column_dimensions[c].width = w
    for code, prof in library.get('profiles', {}).items():
        for kenar, macros in prof.get('operations', {}).items():
            for macro in macros:
                ops = macro.get('ops', [])
                rng_json = json.dumps(macro['ranges'], ensure_ascii=False) if 'ranges' in macro else ''
                for i, op in enumerate(ops):
                    p = op.get('params', {})
                    ws_i.append([code, kenar, macro.get('id',''), macro.get('name',''), i+1,
                                 op.get('p_code','P7'), op.get('tool','T10'), op.get('x_formula','0'),
                                 macro.get('y_value',0), macro.get('z_value',0),
                                 p.get('L',0), p.get('W',0), p.get('C',0), p.get('R',0), p.get('D',0),
                                 rng_json if i == 0 else ''])

    # OzelIslemler
    ws_o = wb.create_sheet('OzelIslemler')
    _hdr(ws_o, ['Op ID','Op Adı','Adım No','P-Kod','Takım','X Formülü','L','W','C','R','D'])
    for w, c in zip([22,26,8,8,8,16,6,6,6,6,6], 'ABCDEFGHIJK'):
        ws_o.column_dimensions[c].width = w
    for op in library.get('custom_ops', []):
        steps = op.get('steps') or [
            {'p_code': op.get('p_code','P7'), 'tool': op.get('tool','T10'),
             'x_formula': xf, 'params': op.get('default_params',{})}
            for xf in op.get('x_formulas', [])]
        for i, s in enumerate(steps):
            p = s.get('params', {})
            ws_o.append([op.get('id',''), op.get('name',''), i+1,
                         s.get('p_code','P7'), s.get('tool','T10'), s.get('x_formula','0'),
                         p.get('L',0), p.get('W',0), p.get('C',0), p.get('R',0), p.get('D',0)])

    os.makedirs(os.path.dirname(fpath), exist_ok=True)
    wb.save(fpath)
    return fpath


def import_from_excel(path: str) -> dict:
    """
    Excel dosyasından kütüphane yükle.
    Dönen dict'i save_library() ile kaydedebilirsiniz.
    """
    openpyxl = _ensure_openpyxl()
    wb  = openpyxl.load_workbook(path, data_only=True)
    lib = _empty_library()

    if 'Profiller' in wb.sheetnames:
        ws = wb['Profiller']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            r = (list(row) + [None]*11)[:11]
            stok_kodu, ad, uretici, seri, tip, kerf, renk, dxf, r_y, r_z, r_v = r
            lib['profiles'][str(stok_kodu)] = {
                'name': str(ad or ''), 'manufacturer': str(uretici or ''),
                'series': str(seri or ''), 'type': str(tip or 'A'),
                'kerf': int(float(kerf or 0)), 'color': str(renk or '#808080'),
                'dxf_file': str(dxf or ''), 'operations': {},
                'robot_y':        int(float(r_y or 400)),
                'robot_z':        int(float(r_z or 400)),
                'robot_vertical': int(float(r_v or 0)),
            }

    if 'Islemler' in wb.sheetnames:
        ws = wb['Islemler']
        macros_map = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            r = (list(row) + [None]*16)[:16]
            (sc, kenar, mid, madi, adim, p_kod, takim, xf, y_mm, z_mm,
             L, W, C, R, D, kosullu_json) = r
            key = (str(sc), str(kenar), str(mid))
            if key not in macros_map:
                m = {'id': str(mid), 'name': str(madi or mid), 'active': True,
                     'y_value': float(y_mm or 0), 'z_value': float(z_mm or 0), 'ops': []}
                if kosullu_json:
                    try: m['ranges'] = json.loads(str(kosullu_json))
                    except Exception: pass
                macros_map[key] = m
            params = {}
            for k, v in zip(_PARAM_KEYS, [L, W, C, R, D]):
                if v is not None:
                    try:
                        fv = float(v)
                        if fv > 0: params[k] = int(fv)
                    except (ValueError, TypeError): pass
            macros_map[key]['ops'].append({
                'label': str(int(float(adim))) if adim else '1',
                'p_code': str(p_kod or 'P7'), 'tool': str(takim or 'T10'),
                'x_formula': str(xf or '0'), 'params': params,
            })
        for (sc, kenar, _), macro in macros_map.items():
            if sc in lib['profiles']:
                (lib['profiles'][sc].setdefault('operations', {})
                 .setdefault(kenar, []).append(macro))

    if 'OzelIslemler' in wb.sheetnames:
        ws = wb['OzelIslemler']
        custom_map = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            r = (list(row) + [None]*11)[:11]
            op_id, op_adi, adim, p_kod, takim, xf, L, W, C, R, D = r
            op_id = str(op_id)
            if op_id not in custom_map:
                custom_map[op_id] = {'id': op_id, 'name': str(op_adi or op_id),
                                      '_custom': True, 'steps': []}
            params = {}
            for k, v in zip(_PARAM_KEYS, [L, W, C, R, D]):
                if v is not None:
                    try:
                        fv = float(v)
                        if fv > 0: params[k] = int(fv)
                    except (ValueError, TypeError): pass
            custom_map[op_id]['steps'].append({
                'p_code': str(p_kod or 'P7'), 'tool': str(takim or 'T10'),
                'x_formula': str(xf or '0'), 'params': params,
            })
        for op in custom_map.values():
            steps = op['steps']
            op['p_code']         = steps[0]['p_code'] if steps else 'P7'
            op['tool']           = steps[0]['tool'] if steps else 'T10'
            op['x_formulas']     = [s['x_formula'] for s in steps]
            op['param_keys']     = list({k for s in steps for k in s['params']})
            op['default_params'] = steps[0]['params'] if steps else {}
        lib['custom_ops'] = list(custom_map.values())

    return lib


# ─────────────────────────────────────────────────────────────────
# Global İşlem Kataloğu — tüm profil türleri için ortak şablon listesi
# Her profil bu katalogdan seçip kendi Y/Z/X değerlerini ayrıca saklar.
# ─────────────────────────────────────────────────────────────────

# ── Toplu Liste: UI'da gösterilen 19 standart işlem ──────────────
BUILTIN_OPS = [
    # 1
    {'id': 'ic_su_tahliye',      'name': 'İç Su Tahliye',       'category': 'Standart',
     'steps': [{'p_code': 'P3', 'tool': 'T40', 'x_formula': '0', 'params': {'L': 25, 'D': 10}}]},
    # 2
    {'id': 'dis_su_tahliye',     'name': 'Dış Su Tahliye',      'category': 'Standart',
     'steps': [{'p_code': 'P3', 'tool': 'T70', 'x_formula': '0', 'params': {'L': 25, 'D': 10}}]},
    # 3
    {'id': 'ic_havalandirma',    'name': 'İç Havalandırma',     'category': 'Standart',
     'steps': [{'p_code': 'P7', 'tool': 'T40', 'x_formula': '0', 'params': {'D': 10}}]},
    # 4
    {'id': 'dis_havalandirma',   'name': 'Dış Havalandırma',    'category': 'Standart',
     'steps': [{'p_code': 'P7', 'tool': 'T70', 'x_formula': '0', 'params': {'D': 10}}]},
    # 5
    {'id': 'mentese_markalama',  'name': 'Menteşe Markalama',   'category': 'Standart',
     'steps': [{'p_code': 'P7', 'tool': 'T30', 'x_formula': '0', 'params': {'D': 2}}]},
    # 6
    {'id': 'mentese_delik',      'name': 'Menteşe Delik',       'category': 'Standart',
     'steps': [{'p_code': 'P7', 'tool': 'T30', 'x_formula': '0', 'params': {'D': 10}}]},
    # 7
    {'id': 'montaj_deligi',      'name': 'Montaj Deliği',       'category': 'Standart',
     'steps': [{'p_code': 'P7', 'tool': 'T50', 'x_formula': '0', 'params': {'D': 10}}]},
    # 8
    {'id': 'uclu_kol_yeri',      'name': 'Üçlü Kol Yeri',      'category': 'Standart',
     'steps': [{'p_code': 'P7', 'tool': 'T71', 'x_formula': '0', 'params': {'D': 35}}]},
    # 9
    {'id': 'ispanyolet_kanali',  'name': 'İspanyolet Kanalı',   'category': 'Standart',
     'steps': [{'p_code': 'P2', 'tool': 'T10', 'x_formula': '0', 'params': {'L': 60, 'W': 12, 'R': 6, 'D': 30}}]},
    # 10
    {'id': 'kilit_kanali',       'name': 'Kilit Kanalı',        'category': 'Standart',
     'steps': [{'p_code': 'P1', 'tool': 'T10', 'x_formula': 'L/2+14', 'params': {'L': 230, 'W': 16, 'D': 30}}]},
    # 11
    {'id': 'ispanyolet_gurubu',  'name': 'İspanyolet Grubu',    'category': 'Standart',
     'steps': [
         {'p_code': 'P2', 'tool': 'T10', 'x_formula': '0', 'y': 20, 'z': 31,
          'params': {'L': 60, 'W': 12, 'R': 6, 'D': 30}},
         {'p_code': 'P7', 'tool': 'T30', 'x_formula': '0', 'y': 35, 'z': 70,
          'params': {'D': 10}},
     ]},
    # 12
    {'id': 'sag_kilit_gurubu',   'name': 'Kilit Grubu Sağ',    'category': 'Standart',
     'steps': [
         {'p_code': 'P6', 'tool': 'T30', 'x_formula': 'L/2',     'y': 35, 'z': 70, 'params': {'C': 18, 'D': 35}},
         {'p_code': 'P6', 'tool': 'T30', 'x_formula': 'L/2',     'y': 35, 'z': 0,  'params': {'C': 18, 'D': 35}},
         {'p_code': 'P5', 'tool': 'T30', 'x_formula': 'L/2-89,5','y': 20, 'z': 70, 'params': {'L': 33, 'W': 10, 'C': 18, 'R': 5, 'D': 35}},
         {'p_code': 'P5', 'tool': 'T70', 'x_formula': 'L/2-89,5','y': 20, 'z': 0,  'params': {'L': 33, 'W': 10, 'C': 18, 'R': 5, 'D': 35}},
         {'p_code': 'P1', 'tool': 'T10', 'x_formula': 'L/2-14',  'y': 20, 'z': 31, 'params': {'L': 230, 'W': 16, 'D': 35}},
     ]},
    # 13
    {'id': 'sol_kilit_gurubu',   'name': 'Kilit Grubu Sol',    'category': 'Standart',
     'steps': [
         {'p_code': 'P6', 'tool': 'T30', 'x_formula': 'L/2',     'y': 35, 'z': 70, 'params': {'C': 18, 'D': 35}},
         {'p_code': 'P6', 'tool': 'T30', 'x_formula': 'L/2',     'y': 35, 'z': 0,  'params': {'C': 18, 'D': 35}},
         {'p_code': 'P4', 'tool': 'T30', 'x_formula': 'L/2+89,5','y': 20, 'z': 70, 'params': {'L': 33, 'W': 10, 'C': 18, 'R': 5, 'D': 35}},
         {'p_code': 'P4', 'tool': 'T70', 'x_formula': 'L/2+89,5','y': 20, 'z': 0,  'params': {'L': 33, 'W': 10, 'C': 18, 'R': 5, 'D': 35}},
         {'p_code': 'P1', 'tool': 'T10', 'x_formula': 'L/2+14',  'y': 20, 'z': 31, 'params': {'L': 230, 'W': 16, 'D': 35}},
     ]},
    # 14
    {'id': 'kol_deligi_ust',     'name': 'Kol Deliği Üst',     'category': 'Standart',
     'steps': [{'p_code': 'P6', 'tool': 'T30', 'x_formula': '0', 'params': {'C': 18, 'D': 30}}]},
    # 15
    {'id': 'kol_deligi_alt',     'name': 'Kol Deliği Alt',     'category': 'Standart',
     'steps': [{'p_code': 'P6', 'tool': 'T70', 'x_formula': '0', 'params': {'C': 18, 'D': 30}}]},
    # 16
    {'id': 'sol_barel_ust',      'name': 'Sol Barel Üst',      'category': 'Standart',
     'steps': [{'p_code': 'P4', 'tool': 'T30', 'x_formula': '0',
                'params': {'L': 33, 'W': 10, 'C': 18, 'R': 5, 'D': 30}}]},
    # 17
    {'id': 'sol_barel_alt',      'name': 'Sol Barel Alt',      'category': 'Standart',
     'steps': [{'p_code': 'P4', 'tool': 'T70', 'x_formula': '0',
                'params': {'L': 33, 'W': 10, 'C': 18, 'R': 5, 'D': 30}}]},
    # 18
    {'id': 'sag_barel_ust',      'name': 'Sağ Barel Üst',     'category': 'Standart',
     'steps': [{'p_code': 'P5', 'tool': 'T30', 'x_formula': '0',
                'params': {'L': 33, 'W': 10, 'C': 18, 'R': 5, 'D': 30}}]},
    # 19
    {'id': 'sag_barel_alt',      'name': 'Sağ Barel Alt',     'category': 'Standart',
     'steps': [{'p_code': 'P5', 'tool': 'T70', 'x_formula': '0',
                'params': {'L': 33, 'W': 10, 'C': 18, 'R': 5, 'D': 30}}]},
]

# ── Eski ID'ler — UI'da görünmez, eski profil verisi için geriye dönük uyumluluk ──
_LEGACY_OPS = [
    {'id': 'ic_su_tahliye_kasa',      'name': 'İç Su Tahliye (Kasa)',
     'steps': [{'p_code': 'P3', 'tool': 'T40', 'x_formula': '0', 'params': {'L': 25, 'D': 10}}]},
    {'id': 'dis_su_tahliye_kasa',     'name': 'Dış Su Tahliye (Kasa)',
     'steps': [{'p_code': 'P3', 'tool': 'T70', 'x_formula': '0', 'params': {'L': 25, 'D': 10}}]},
    {'id': 'ic_su_tahliye_kanat',     'name': 'İç Su Tahliye (Kanat)',
     'steps': [{'p_code': 'P3', 'tool': 'T60', 'x_formula': '0', 'params': {'L': 25, 'D': 10}}]},
    {'id': 'dis_su_tahliye_kanat',    'name': 'Dış Su Tahliye (Kanat)',
     'steps': [{'p_code': 'P3', 'tool': 'T10', 'x_formula': '0', 'params': {'L': 25, 'D': 10}}]},
    {'id': 'ic_havalandirma_kasa',    'name': 'İç Havalandırma (Kasa)',
     'steps': [{'p_code': 'P7', 'tool': 'T40', 'x_formula': '0', 'params': {'D': 10}}]},
    {'id': 'dis_havalandirma_kasa',   'name': 'Dış Havalandırma (Kasa)',
     'steps': [{'p_code': 'P7', 'tool': 'T70', 'x_formula': '0', 'params': {'D': 10}}]},
    {'id': 'ic_havalandirma_kanat',   'name': 'İç Havalandırma (Kanat)',
     'steps': [{'p_code': 'P7', 'tool': 'T60', 'x_formula': '0', 'params': {'D': 10}}]},
    {'id': 'dis_havalandirma_kanat',  'name': 'Dış Havalandırma (Kanat)',
     'steps': [{'p_code': 'P7', 'tool': 'T10', 'x_formula': '0', 'params': {'D': 10}}]},
    {'id': 'mentese_markalama_kanat', 'name': 'Menteşe Markalama (Kanat)',
     'steps': [{'p_code': 'P7', 'tool': 'T50', 'x_formula': '0', 'params': {'D': 2}}]},
    {'id': 'karsilik_markalama',      'name': 'Karşılık Markalama',
     'steps': [{'p_code': 'P7', 'tool': 'T50', 'x_formula': '0', 'params': {'D': 2}}]},
    {'id': 'ust_mentese',             'name': 'Üst Menteşe Yeri',
     'steps': [{'p_code': 'P7', 'tool': 'T30', 'x_formula': '0', 'params': {'D': 10}}]},
    {'id': 'alt_mentese',             'name': 'Alt Menteşe Yeri',
     'steps': [{'p_code': 'P7', 'tool': 'T30', 'x_formula': '0', 'params': {'D': 10}}]},
    {'id': 'orta_mentese',            'name': 'Orta Menteşe Yeri',
     'steps': [{'p_code': 'P7', 'tool': 'T30', 'x_formula': 'L/2', 'params': {'D': 10}}]},
    {'id': 'kilit_ust_kanal',         'name': 'Kilit Üst Kanal',
     'steps': [{'p_code': 'P1', 'tool': 'T10', 'x_formula': '0', 'params': {'L': 150, 'W': 16, 'D': 30}}]},
    {'id': 'kilit_alt_kanal',         'name': 'Kilit Alt Kanal',
     'steps': [{'p_code': 'P1', 'tool': 'T10', 'x_formula': '0', 'params': {'L': 150, 'W': 16, 'D': 30}}]},
    {'id': 'ust_kilit_grubu',         'name': 'Üst Kilit Grubu',
     'steps': [
         {'p_code': 'P6', 'tool': 'T30', 'x_formula': 'L/2',  'y': 35, 'z': 70, 'params': {'C': 18, 'D': 35}},
         {'p_code': 'P6', 'tool': 'T30', 'x_formula': 'L/2',  'y': 35, 'z': 0,  'params': {'C': 18, 'D': 35}},
         {'p_code': 'P4', 'tool': 'T30', 'x_formula': '0',    'y': 20, 'z': 70, 'params': {'L': 33, 'W': 10, 'C': 18, 'R': 5, 'D': 35}},
         {'p_code': 'P4', 'tool': 'T70', 'x_formula': '0',    'y': 20, 'z': 0,  'params': {'L': 33, 'W': 10, 'C': 18, 'R': 5, 'D': 35}},
     ]},
    {'id': 'alt_kilit_grubu',         'name': 'Alt Kilit Grubu',
     'steps': [
         {'p_code': 'P6', 'tool': 'T30', 'x_formula': 'L/2',  'y': 35, 'z': 70, 'params': {'C': 18, 'D': 35}},
         {'p_code': 'P6', 'tool': 'T30', 'x_formula': 'L/2',  'y': 35, 'z': 0,  'params': {'C': 18, 'D': 35}},
         {'p_code': 'P5', 'tool': 'T30', 'x_formula': '0',    'y': 20, 'z': 70, 'params': {'L': 33, 'W': 10, 'C': 18, 'R': 5, 'D': 35}},
         {'p_code': 'P5', 'tool': 'T70', 'x_formula': '0',    'y': 20, 'z': 0,  'params': {'L': 33, 'W': 10, 'C': 18, 'R': 5, 'D': 35}},
     ]},
    {'id': 'tek_delik_kucuk',         'name': 'Tek Delik (Küçük)',
     'steps': [{'p_code': 'P7', 'tool': 'T30', 'x_formula': '0', 'params': {'D': 10}}]},
    {'id': 'tek_delik_orta',          'name': 'Tek Delik (Orta)',
     'steps': [{'p_code': 'P7', 'tool': 'T30', 'x_formula': '0', 'params': {'D': 16}}]},
    {'id': 'tek_delik_buyuk',         'name': 'Tek Delik (Büyük)',
     'steps': [{'p_code': 'P7', 'tool': 'T50', 'x_formula': '0', 'params': {'D': 22}}]},
    {'id': 'cift_delik',              'name': 'Çift Delik',
     'steps': [
         {'p_code': 'P7', 'tool': 'T30', 'x_formula': '0', 'params': {'D': 10}},
         {'p_code': 'P7', 'tool': 'T30', 'x_formula': '0', 'params': {'D': 10}},
     ]},
    {'id': 'dizi_delik',              'name': 'Dizi Delik',
     'steps': [
         {'p_code': 'P7', 'tool': 'T30', 'x_formula': '0', 'params': {'D': 10}},
         {'p_code': 'P7', 'tool': 'T30', 'x_formula': '0', 'params': {'D': 10}},
         {'p_code': 'P7', 'tool': 'T30', 'x_formula': '0', 'params': {'D': 10}},
     ]},
    {'id': 'freze_kanalı_ust',        'name': 'Freze Kanalı Üst',
     'steps': [{'p_code': 'P2', 'tool': 'T40', 'x_formula': '0', 'params': {'L': 100, 'W': 10, 'R': 5, 'D': 15}}]},
    {'id': 'freze_kanali_alt',        'name': 'Freze Kanalı Alt',
     'steps': [{'p_code': 'P2', 'tool': 'T70', 'x_formula': '0', 'params': {'L': 100, 'W': 10, 'R': 5, 'D': 15}}]},
    {'id': 'freze_cep',               'name': 'Freze Cep',
     'steps': [{'p_code': 'P2', 'tool': 'T40', 'x_formula': '0', 'params': {'L': 60, 'W': 20, 'R': 10, 'D': 20}}]},
    {'id': 'robot_kesim',             'name': 'Robot Kesim',
     'steps': [{'p_code': 'P7', 'tool': 'T10', 'x_formula': '0', 'params': {'D': 0}}]},
    {'id': 'robot_markalama',         'name': 'Robot Markalama',
     'steps': [{'p_code': 'P7', 'tool': 'T30', 'x_formula': '0', 'params': {'D': 0}}]},
]

# Sabit BUILTIN ID kümesi (hız için — sadece görünür 19 işlem)
_BUILTIN_IDS = {op['id'] for op in BUILTIN_OPS}
# Tüm bilinen ID'ler (legacy dahil)
_ALL_KNOWN_IDS = _BUILTIN_IDS | {op['id'] for op in _LEGACY_OPS}


def get_catalog_ops(library: dict) -> list:
    """
    Tüm profiller için ortak global işlem kataloğunu döndürür.
    = BUILTIN_OPS  +  library['custom_ops']
    Aynı ID'ye sahip custom op, builtin olanı gizler (kullanıcı override'ı).
    """
    custom = library.get('custom_ops', [])
    custom_ids = {op['id'] for op in custom}
    result = [op for op in BUILTIN_OPS if op['id'] not in custom_ids]
    result.extend(custom)
    return result


def get_catalog_op(library: dict, op_id: str):
    """
    Katalogdan ID'ye göre tek bir işlem şablonu döndürür.
    Önce görünür katalog (19 işlem + custom), bulunamazsa legacy ops'a bakar.
    """
    for op in get_catalog_ops(library):
        if op['id'] == op_id:
            return op
    # Eski profil verisi için legacy arama
    for op in _LEGACY_OPS:
        if op['id'] == op_id:
            return op
    return None


# ─────────────────────────────────────────────────────────────────
# Ozel islem yardimcilari
# ─────────────────────────────────────────────────────────────────

def get_custom_ops(library: dict) -> list:
    """Kullanici tarafindan eklenen ozel islemler listesi."""
    return library.get('custom_ops', [])


def add_custom_op(library: dict, op: dict) -> bool:
    """Yeni ozel islem ekle. ID cakisiyorsa False doner."""
    ops = library.setdefault('custom_ops', [])
    if any(o['id'] == op['id'] for o in ops):
        return False
    ops.append(op)
    return True


def delete_custom_op(library: dict, op_id: str):
    """Ozel islemi kalici olarak sil."""
    library['custom_ops'] = [
        o for o in library.get('custom_ops', []) if o['id'] != op_id
    ]


# ─────────────────────────────────────────────────────────────────
# Sorgu yardımcıları
# ─────────────────────────────────────────────────────────────────

def get_profiles_by_type(library: dict, ptype: str) -> list:
    result = []
    for code, prof in library.get('profiles', {}).items():
        if prof.get('type', '').upper() == ptype.upper():
            entry = dict(prof)
            entry['stock_code'] = code
            result.append(entry)
    result.sort(key=lambda p: p.get('name', ''))
    return result


def get_all_types_in_library(library: dict) -> list:
    types = set()
    for prof in library.get('profiles', {}).values():
        t = prof.get('type', '')
        if t:
            types.add(t.upper())
    return sorted(types)


def get_last_used(library: dict, ptype: str) -> str:
    return library.get('last_used', {}).get(ptype.upper(), '')


def set_last_used(library: dict, ptype: str, stock_code: str):
    library.setdefault('last_used', {})[ptype.upper()] = stock_code


def get_profile(library: dict, stock_code: str):
    return library.get('profiles', {}).get(stock_code)


def get_tools(library: dict) -> dict:
    return library.get('tools', {})


# ─────────────────────────────────────────────────────────────────
# P-code üretim motoru
# ─────────────────────────────────────────────────────────────────

def _get_ops_for_length(macro: dict, length_x10: int) -> list:
    """
    Makro için length_x10'a göre doğru ops listesini döndürür.
    Eğer makroda 'ranges' varsa, uygun aralığın x_formulas'ından ops üretir.
    Yoksa normal 'ops' listesini döndürür (geriye dönük uyumluluk).

    ranges formatı:
      [{'max_mm': 500, 'x_formulas': ['L/2']},
       {'max_mm': 1000, 'x_formulas': ['L-300', '300']},
       {'max_mm': None, 'x_formulas': ['L-1500', '1500']}]
    max_mm=None → sınırsız (en son aralık olarak kullanılır).
    """
    ranges = macro.get('ranges')
    if not ranges:
        return macro.get('ops', [])

    length_mm = length_x10 / 10.0
    # Şablon op: p_code, tool, params için ilk mevcut ops kaydını kullan
    base_ops = macro.get('ops', [])
    template = base_ops[0] if base_ops else {'p_code': 'P7', 'tool': 'T10', 'params': {}}

    # Aralıkları küçükten büyüğe sırala (None=sınırsız sona gitsin)
    sorted_ranges = sorted(ranges, key=lambda r: r.get('max_mm') or float('inf'))

    for rng in sorted_ranges:
        max_mm = rng.get('max_mm')
        if max_mm is None or length_mm <= max_mm:
            # Bu aralık eşleşti — x_formulas'tan ops üret
            return [
                {**template, 'x_formula': xf, 'label': str(i + 1)}
                for i, xf in enumerate(rng.get('x_formulas', []))
            ]

    # Hiçbir aralık uymadıysa son aralığı kullan
    last = sorted_ranges[-1]
    return [
        {**template, 'x_formula': xf, 'label': str(i + 1)}
        for i, xf in enumerate(last.get('x_formulas', []))
    ]


def generate_side_code(stock_code: str, length_x10: int,
                       side: str, library: dict,
                       active_only: bool = True,
                       x0_mode: bool = False) -> str:
    """
    Belirli profil + kenar için tüm makroların P-code string'ini üretir.
    length_x10 : profil uzunluğu ×10 cinsinden (1200mm → 12000)
    side        : 'ALT', 'ÜST', 'SOL', 'SAĞ'
    x0_mode     : True ise tüm X değerleri 0 olarak üretilir (test/taslak)
    """
    profile = get_profile(library, stock_code)
    if not profile:
        return ''
    macros = profile.get('operations', {}).get(side, [])
    result = ''
    for macro in macros:
        if active_only and not macro.get('active', True):
            continue
        y_mm = macro.get('y_value', 0)
        z_mm = macro.get('z_value', 0)
        # Ranges desteği: uzunluğa göre doğru ops listesini seç
        ops_list = _get_ops_for_length(macro, length_x10)
        for op_def in ops_list:
            if x0_mode:
                test_op = dict(op_def)
                test_op['x_formula'] = '0'
                result += _build_op(test_op, length_x10, y_mm, z_mm)
            else:
                result += _build_op(op_def, length_x10, y_mm, z_mm)
    return result


def get_side_op_names(stock_code: str, side: str, library: dict,
                      active_only: bool = True) -> list[str]:
    """
    Belirli profil + kenar için aktif makroların isim listesini döndürür.
    Örn: ['İç Su Tahliye', 'Menteşe Markalama', 'Montaj Deliği']
    """
    profile = get_profile(library, stock_code)
    if not profile:
        return []
    macros = profile.get('operations', {}).get(side, [])
    names = []
    for macro in macros:
        if active_only and not macro.get('active', True):
            continue
        name = macro.get('name', macro.get('id', '?'))
        if name:
            # Kümülatif 🔗 grup işaretlerini temizle (eski kayıtlardan birikmişse)
            # replace() kullan: "🔗 🔗 🔗 Ad" gibi aralıklı durumları da temizler
            name = name.replace('\U0001f517', '').strip()
            if name:
                names.append(name)
    return names


def generate_frame_codes(frame_pieces: list, library: dict) -> list:
    """
    Çerçeve parça listesi için P-code'ları hesaplar.
    Her parça dict'e 'generated_code' alanı ekler.
    """
    result = []
    for piece in frame_pieces:
        p = dict(piece)
        code = generate_side_code(
            p.get('stock_code', ''),
            p.get('length', 0),
            p.get('side', ''),
            library,
        )
        p['generated_code'] = code
        result.append(p)
    return result


# ─────────────────────────────────────────────────────────────────
# İç: tek işlem kodu oluşturma
# ─────────────────────────────────────────────────────────────────

def _calc_x(formula: str, L_mm: float) -> float:
    """
    X formülünü değerlendirir.
    L_mm  : parça uzunluğu milimetre cinsinden (örn. 1500.0)
    Dönüş : sonuç mm cinsinden float (örn. L/2+89,5 → 644.5)
    Not   : _build_op bu değeri ×10 ederek MDB formatına çevirir.

    ⚠ Önemli — X=0 neden çıkar?
      • "L-1500" formülü, L=1500mm profil için 1500-1500=0 → X0 üretir.
        Bu DOĞRU davranıştır: X=0 profilin başlangıç noktasıdır.
      • Farklı uzunluklarda (ör. L=2000mm): 2000-1500=500 → X5000 üretir.
      • İşlem tanımlama ekranındaki "Test uzunluğu" alanı ile
        formülün farklı uzunluklarda ne ürettiğini canlı görebilirsiniz.
    CNC makineleri negatif X konumu kabul etmez; negatif sonuçlar
    max(0.0,...) ile sıfıra sabitlenir.
    """
    try:
        # Türkçe ondalık virgülünü noktaya çevir (örn: "+89,5" → "+89.5")
        f = formula.strip().replace(',', '.').replace('L', str(L_mm))
        result = eval(f, {'__builtins__': {}})
        return max(0.0, float(result))
    except Exception:
        return 0.0


def _x10(params: dict, key: str) -> int:
    try:
        return int(float(params.get(key, 0)) * 10)
    except (TypeError, ValueError):
        return 0


def _build_op(op_def: dict, length_x10: int, y_mm: float, z_mm: float) -> str:
    op   = op_def.get('p_code', 'P7')
    tool = op_def.get('tool', 'T10')
    prm  = op_def.get('params', {})

    # Formüller mm cinsinden L kullanır → sonucu ×10 yaparak MDB formatına çevir
    length_mm = length_x10 / 10.0
    x = int(round(_calc_x(op_def.get('x_formula', '0'), length_mm) * 10))
    # Adıma özgü y/z varsa makro seviyesini geçersiz kıl
    step_y = op_def.get('y')
    step_z = op_def.get('z')
    y = int(round(float(step_y) * 10)) if step_y is not None else int(round(y_mm * 10))
    z = int(round(float(step_z) * 10)) if step_z is not None else int(round(z_mm * 10))

    base = f'{op}{tool}X{x}Y{y}Z{z}'

    if op == 'P7':
        return f'{base}D{_x10(prm,"D")}//'
    if op == 'P3':
        if 'L' in prm:
            return f'{base}L{_x10(prm,"L")}D{_x10(prm,"D")}//'
        return f'{base}D{_x10(prm,"D")}//'
    if op == 'P2':
        L_v = _x10(prm,'L'); W_v = _x10(prm,'W')
        R_v = _x10(prm,'R') if 'R' in prm else W_v // 2
        return f'{base}L{L_v}W{W_v}R{R_v}D{_x10(prm,"D")}//'
    if op == 'P1':
        return f'{base}L{_x10(prm,"L")}W{_x10(prm,"W")}D{_x10(prm,"D")}//'
    if op in ('P4','P5'):
        L_v=_x10(prm,'L'); W_v=_x10(prm,'W'); C_v=_x10(prm,'C'); D_v=_x10(prm,'D')
        R_v=_x10(prm,'R') if 'R' in prm else W_v//2
        return f'{base}L{L_v}W{W_v}C{C_v}R{R_v}D{D_v}//'
    if op == 'P6':
        return f'{base}C{_x10(prm,"C")}D{_x10(prm,"D")}//'
    return f'{base}//'


# ─────────────────────────────────────────────────────────────────
# Kütüphane CRUD yardımcıları
# ─────────────────────────────────────────────────────────────────

def add_profile(library: dict, stock_code: str, profile: dict) -> bool:
    if stock_code in library.get('profiles', {}):
        return False
    library.setdefault('profiles', {})[stock_code] = profile
    return True


def delete_profile(library: dict, stock_code: str) -> bool:
    if stock_code not in library.get('profiles', {}):
        return False
    del library['profiles'][stock_code]
    for k, v in library.get('last_used', {}).items():
        if v == stock_code:
            library['last_used'][k] = ''
    return True


def duplicate_profile(library: dict, src_code: str, new_code: str, new_name: str) -> bool:
    src = get_profile(library, src_code)
    if not src or new_code in library.get('profiles', {}):
        return False
    new_prof = copy.deepcopy(src)
    new_prof['name'] = new_name
    library['profiles'][new_code] = new_prof
    return True


def new_empty_macro(macro_id: str = 'yeni_islem', name: str = 'Yeni İşlem') -> dict:
    return {
        'id': macro_id, 'name': name, 'active': True,
        'y_value': 0, 'z_value': 0,
        'ops': [{'label':'','p_code':'P7','tool':'T10','x_formula':'L/2','params':{'D':8}}],
    }


def new_empty_profile(stock_code: str = '', ptype: str = 'A') -> dict:
    return {
        'type': ptype, 'name': stock_code, 'manufacturer': '',
        'series': '', 'dxf_file': '', 'kerf': 45, 'color': '#808080',
        'operations': {s: [] for s in SIDES},
    }
